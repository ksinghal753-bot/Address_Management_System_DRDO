import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def export_to_excel(records: list[dict], output_path: str, columns: list[tuple[str, str]]) -> tuple[bool, str]:
    """
    Export a list of records to an Excel (.xlsx) file.
    `columns` should be a list of tuples: (header_name, dict_key).
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Addresses"

        # 1. Write Headers
        header_font = Font(bold=True)
        for col_idx, (header_name, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2. Write Data
        for row_idx, record in enumerate(records, start=2):
            for col_idx, (_, dict_key) in enumerate(columns, start=1):
                val = record.get(dict_key, "")
                if val is None:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                # Set as text format to prevent Excel from altering pin codes or dates
                cell.number_format = '@'

        # 3. Auto-adjust column widths
        for col_idx, (header_name, _) in enumerate(columns, start=1):
            max_length = len(header_name)
            for row_idx in range(2, len(records) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_length = max(max_length, len(str(cell_val)))
            
            # Add a little padding
            adjusted_width = max_length + 2
            # Cap at 50 to avoid ridiculously wide columns
            adjusted_width = min(adjusted_width, 50)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(output_path)
        return True, "Export successful."
    except Exception as e:
        return False, f"Failed to export Excel: {e}"
